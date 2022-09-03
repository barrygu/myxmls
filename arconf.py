#!/usr/bin/python

__version_info__ = ('1', '0', '0')
__version__ = '.'.join(__version_info__)
__author__ = 'Gu Jianzhong <jianzhong.gu@harman.com>'

import os, sys

prog = os.path.basename(sys.argv[0])
argc = len(sys.argv)
if argc != 2:
    print("\nInvalid argument.\nUsage: ", file=sys.stderr)
    print(f"   {prog} <excel_file>", file=sys.stderr)
    sys.exit(1)
infile = sys.argv[1]

try:
    import xml.etree.cElementTree as e
except ImportError:
    import xml.etree.ElementTree as e

root = e.Element("audioRoutingConfiguration")
root.set("version", "1.0")
root.set("xmlns:xi", "http://www.w3.org/2001/XInclude")

elenames={
    "audioswitch":"switch",
    "devicesconf":"device",
    "alsaportsconf":"port",
    "moduleslist":"module",
    "pathconf":"path"
}

def V(s):
    module, chmap = (s.split("%") + [""])[0:2]
    mtype, name = module.split(":")
    mtype = mtype.replace("[","").replace("]","").strip()
    return ("[" + mtype + "]:" + name.strip(), chmap.strip())

import openpyxl
wb = openpyxl.load_workbook(infile,data_only=True,read_only=True)
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    max_col = ws.max_column
    rows = list(ws.rows)
    headers = []
    for cell in rows[0]:
        if cell.value is None: break
        headers.append(cell.value.strip())
    max_col = len(headers)
    #print("%s: %s" % (sheetname, headers))
    t = e.SubElement(root,sheetname)

    for row in rows[1:]:
        if row[0].value is None: break
        s = e.SubElement(t,elenames[sheetname])
        nvs = dict(zip(headers, [str(cell.value) if cell.value or cell.value == 0 else "" for cell in row][:max_col]))
        for hd in headers:
            if hd == "path":
                p = nvs[hd].replace("\n", "").split(";")
                c = e.SubElement(s, "connections")
                for it in p:
                    cc = e.SubElement(c, "connection")
                    left,right = it.split("-->")
                    a,b = V(left)
                    cc.set("from",a)
                    cc.set("outchmap",b)
                    a,b = V(right)
                    cc.set("to",a)
                    cc.set("inchmap",b)
            elif hd == "periodtime":
                pass
            else:
                s.set(hd, nvs[hd].strip())
t = e.SubElement(root,"xi:include")
t.set("href", "audio_processing_configuration.xml")
wb.close()

a = e.ElementTree(root)
e.indent(a)
#a.write("arconf.xml")
#with open("arconf.xml", "wb") as f:
f=sys.stdout.buffer
f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'.encode('utf8'))
f.write('<?xml-stylesheet type="text/xsl" href="arconf.xsl"?>\n'.encode('utf8'))
a.write(f, "utf-8")
